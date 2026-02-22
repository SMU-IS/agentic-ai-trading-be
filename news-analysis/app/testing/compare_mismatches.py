"""
compare_mismatches.py

Compares sentiment mismatches between Qwen and Llama.
Only includes posts that appear in BOTH mismatch files (i.e., both models
flagged at least one sentiment error on that post).

For each such post, shows all tickers that either model got wrong, side by side,
along with ground truth from curated_dataset_200.json — for manual review.

Usage:
    python compare_mismatches.py

Output:
    sentiment_analysis/sentiment_mismatch_comparison.xlsx

Requirements:
    pip install openpyxl
"""

import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install it with:  pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
TESTING_DIR = os.path.dirname(os.path.abspath(__file__))
SENTIMENT_DIR = os.path.join(TESTING_DIR, "sentiment_analysis", "batch_output")

QWEN_MISMATCH_PATH  = os.path.join(SENTIMENT_DIR, "Qwen",  "sentiment_Qwen_mismatches.json")
LLAMA_MISMATCH_PATH = os.path.join(SENTIMENT_DIR, "Llama", "sentiment_Llama_mismatches.json")
CURATED_DATASET_PATH = os.path.join(TESTING_DIR, "curated_dataset_200.json")

OUTPUT_PATH = os.path.join(TESTING_DIR, "sentiment_analysis", "sentiment_mismatch_comparison.xlsx")

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL     = PatternFill("solid", fgColor="1F3864")   # dark navy
POST_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")  # medium blue (post separator)
GT_FILL         = PatternFill("solid", fgColor="E2EFDA")   # pale green (ground truth col)
BOTH_WRONG_FILL = PatternFill("solid", fgColor="FFE699")   # amber  (both models wrong)
ONE_WRONG_FILL  = PatternFill("solid", fgColor="F2F2F2")   # light grey (one model wrong)

SENTIMENT_FILLS = {
    "positive": PatternFill("solid", fgColor="C6EFCE"),    # green
    "negative": PatternFill("solid", fgColor="FFC7CE"),    # red
    "neutral":  PatternFill("solid", fgColor="FFEB9C"),    # yellow
}
NONE_FILL = PatternFill("solid", fgColor="D9D9D9")         # grey (None/skipped)

WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
BOLD_FONT   = Font(bold=True, name="Calibri", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
ITALIC_FONT = Font(italic=True, name="Calibri", size=10)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
THICK_TOP_BORDER = Border(
    left=Side(style="thin"),   right=Side(style="thin"),
    top=Side(style="medium"),  bottom=Side(style="thin"),
)

TOP_WRAP = Alignment(wrap_text=True, vertical="top")
CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _build_mismatch_map(mismatch_list: list) -> dict:
    """Returns {post_id: {ticker: mismatch_data}} from a mismatch JSON list."""
    result = {}
    for entry in mismatch_list:
        post_id = entry["id"]
        result[post_id] = {m["ticker"]: m for m in entry.get("ticker_mismatches", [])}
    return result


def _build_curated_map(curated_list: list) -> dict:
    """Returns {post_id: post_dict} from the curated dataset."""
    return {p["id"]: p for p in curated_list}


def _sentiment_fill(label):
    if label in SENTIMENT_FILLS:
        return SENTIMENT_FILLS[label]
    return NONE_FILL


def _write_cell(ws, row: int, col: int, value,
                font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font      or NORMAL_FONT
    cell.fill      = fill      or PatternFill()          # no fill by default
    cell.alignment = alignment or TOP_WRAP
    cell.border    = border    or THIN_BORDER
    return cell


def _merge_and_style(ws, start_row: int, end_row: int, col: int,
                     value, font=None, fill=None, border=None):
    """Write value in start_row, merge down to end_row, apply style."""
    cell = ws.cell(row=start_row, column=col, value=value)
    cell.font      = font   or NORMAL_FONT
    cell.fill      = fill   or PatternFill()
    cell.alignment = TOP_WRAP
    cell.border    = border or THIN_BORDER
    if end_row > start_row:
        ws.merge_cells(
            start_row=start_row, start_column=col,
            end_row=end_row,     end_column=col,
        )
        # Re-apply alignment after merge (openpyxl requirement)
        ws.cell(row=start_row, column=col).alignment = TOP_WRAP


# ── Column definitions ────────────────────────────────────────────────────────

COLUMNS = [
    # (header,                    width,  group)
    ("Post ID",                    22,    "post"),
    ("Post URL",                   35,    "post"),
    ("Subreddit",                  14,    "post"),
    ("Post Text",                  60,    "post"),
    ("Ticker",                     10,    "ticker"),
    ("GT Sentiment",               14,    "gt"),
    ("GT Event Type",              22,    "gt"),
    ("GT Event Proposal",          30,    "gt"),
    ("Qwen Label",                 14,    "qwen"),
    ("Qwen Score",                 10,    "qwen"),
    ("Qwen Reasoning",             48,    "qwen"),
    ("Qwen Mismatch?",             14,    "qwen"),
    ("Llama Label",                14,    "llama"),
    ("Llama Score",                10,    "llama"),
    ("Llama Reasoning",            48,    "llama"),
    ("Llama Mismatch?",            14,    "llama"),
    ("Both Models Wrong?",         16,    "both"),
    ("Verdict / Notes",            32,    "review"),
]

# Column indices (1-based) by name — built from COLUMNS list
COL = {h: i + 1 for i, (h, _, _) in enumerate(COLUMNS)}


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    qwen_mismatches  = _load_json(QWEN_MISMATCH_PATH)
    llama_mismatches = _load_json(LLAMA_MISMATCH_PATH)
    curated          = _load_json(CURATED_DATASET_PATH)

    qwen_map    = _build_mismatch_map(qwen_mismatches)
    llama_map   = _build_mismatch_map(llama_mismatches)
    curated_map = _build_curated_map(curated)

    # Posts in BOTH mismatch files (both models had at least one error here)
    both_ids = sorted(set(qwen_map.keys()) & set(llama_map.keys()))

    qwen_only  = set(qwen_map)  - set(llama_map)
    llama_only = set(llama_map) - set(qwen_map)

    print(f"  Posts in Qwen mismatches : {len(qwen_map)}")
    print(f"  Posts in Llama mismatches: {len(llama_map)}")
    print(f"  Posts in BOTH            : {len(both_ids)}  <-- written to Excel")
    print(f"  Posts only in Qwen       : {len(qwen_only)}")
    print(f"  Posts only in Llama      : {len(llama_only)}")

    # ── Workbook setup ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Both Mismatched"
    ws.freeze_panes = "E2"   # freeze post-level columns + header

    # Header row
    for i, (header, width, _) in enumerate(COLUMNS):
        col = i + 1
        c = ws.cell(row=1, column=col, value=header)
        c.font      = WHITE_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_row = 2
    total_ticker_rows = 0

    for post_id in both_ids:
        post         = curated_map.get(post_id, {})
        content      = post.get("content", {})
        post_text    = content.get("clean_combined_withurl", "")[:1500]
        subreddit    = post.get("metadata", {}).get("subreddit", "")
        post_url     = post.get("url", "")
        correct_meta = post.get("correct_metadata", {})

        qwen_tickers  = set(qwen_map[post_id].keys())
        llama_tickers = set(llama_map[post_id].keys())
        all_tickers   = sorted(qwen_tickers | llama_tickers)

        post_start = current_row

        for ticker in all_tickers:
            # Ground truth for this ticker
            gt_meta     = correct_meta.get(ticker, {})
            gt_sentiment = gt_meta.get("sentiment_label", "N/A")
            gt_event     = gt_meta.get("event_type", "N/A")
            gt_proposal_raw = gt_meta.get("event_proposal")
            if isinstance(gt_proposal_raw, dict):
                gt_proposal = json.dumps(gt_proposal_raw, ensure_ascii=False)
            else:
                gt_proposal = gt_proposal_raw or ""

            # Qwen data
            qd             = qwen_map[post_id].get(ticker, {})
            qwen_label     = qd.get("generated_label")   # None if not in Qwen mismatches
            qwen_score     = qd.get("generated_score")
            qwen_reasoning = qd.get("generated_reasoning", "")
            qwen_mismatch  = ticker in qwen_tickers

            # Llama data
            ld              = llama_map[post_id].get(ticker, {})
            llama_label     = ld.get("generated_label")
            llama_score     = ld.get("generated_score")
            llama_reasoning = ld.get("generated_reasoning", "")
            llama_mismatch  = ticker in llama_tickers

            both_wrong = qwen_mismatch and llama_mismatch

            # Row background: amber if both wrong, light grey if only one wrong
            row_fill = BOTH_WRONG_FILL if both_wrong else ONE_WRONG_FILL

            # Border: slightly thicker top for first ticker row of each post
            top_border = THICK_TOP_BORDER if (current_row == post_start) else THIN_BORDER

            def wc(col_name, value, font=None, fill=None):
                """Shortcut: write cell in current_row at named column."""
                f = fill if fill is not None else row_fill
                _write_cell(
                    ws, current_row, COL[col_name], value,
                    font=font or NORMAL_FONT,
                    fill=f,
                    border=top_border if current_row == post_start else THIN_BORDER,
                )

            # Ticker and GT columns (not merged — one row per ticker)
            wc("Ticker",           ticker,          font=BOLD_FONT)
            wc("GT Sentiment",     gt_sentiment,    fill=_sentiment_fill(gt_sentiment))
            wc("GT Event Type",    gt_event,        fill=GT_FILL)
            wc("GT Event Proposal", gt_proposal,    fill=GT_FILL)

            # Qwen columns
            ql_fill = _sentiment_fill(qwen_label) if qwen_mismatch else NONE_FILL
            wc("Qwen Label",     qwen_label     if qwen_mismatch else "–", fill=ql_fill)
            wc("Qwen Score",     qwen_score     if qwen_mismatch else "–")
            wc("Qwen Reasoning", qwen_reasoning if qwen_mismatch else "–",
               font=ITALIC_FONT if not qwen_mismatch else NORMAL_FONT)
            wc("Qwen Mismatch?", "YES" if qwen_mismatch else "no",
               font=BOLD_FONT if qwen_mismatch else NORMAL_FONT)

            # Llama columns
            ll_fill = _sentiment_fill(llama_label) if llama_mismatch else NONE_FILL
            wc("Llama Label",     llama_label     if llama_mismatch else "–", fill=ll_fill)
            wc("Llama Score",     llama_score     if llama_mismatch else "–")
            wc("Llama Reasoning", llama_reasoning if llama_mismatch else "–",
               font=ITALIC_FONT if not llama_mismatch else NORMAL_FONT)
            wc("Llama Mismatch?", "YES" if llama_mismatch else "no",
               font=BOLD_FONT if llama_mismatch else NORMAL_FONT)

            # Summary columns
            wc("Both Models Wrong?",
               "YES" if both_wrong else "no",
               font=BOLD_FONT if both_wrong else NORMAL_FONT)
            wc("Verdict / Notes", "")   # blank for reviewer

            ws.row_dimensions[current_row].height = 80
            current_row += 1
            total_ticker_rows += 1

        post_end = current_row - 1

        # Merge post-level cells (Post ID, URL, Subreddit, Post Text) across ticker rows
        post_fill = PatternFill()   # no extra fill for merged cells (let row colour show)
        for col_name, value in [
            ("Post ID",   post_id),
            ("Post URL",  post_url),
            ("Subreddit", subreddit),
            ("Post Text", post_text),
        ]:
            _merge_and_style(
                ws, post_start, post_end, COL[col_name],
                value=value,
                font=BOLD_FONT if col_name == "Post ID" else NORMAL_FONT,
                border=THICK_TOP_BORDER,
            )

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12

    summary_rows = [
        ("Metric",                              "Count"),
        ("Posts in Qwen mismatches",            len(qwen_map)),
        ("Posts in Llama mismatches",           len(llama_map)),
        ("Posts in BOTH (written to main sheet)", len(both_ids)),
        ("Posts only in Qwen mismatches",       len(qwen_only)),
        ("Posts only in Llama mismatches",      len(llama_only)),
        ("Total ticker rows in Excel",          total_ticker_rows),
    ]

    for r, (label, val) in enumerate(summary_rows, start=1):
        is_header = r == 1
        lc = ws2.cell(row=r, column=1, value=label)
        vc = ws2.cell(row=r, column=2, value=val)
        for c in (lc, vc):
            c.font      = WHITE_FONT if is_header else BOLD_FONT if r == 4 else NORMAL_FONT
            c.fill      = HEADER_FILL if is_header else BOTH_WRONG_FILL if r == 4 else PatternFill()
            c.alignment = CENTER if is_header else TOP_WRAP
            c.border    = THIN_BORDER

    # Colour legend
    ws2.cell(row=len(summary_rows) + 2, column=1, value="Colour Legend").font = BOLD_FONT
    legend = [
        ("Amber row",        "Both Qwen AND Llama got this ticker wrong",          BOTH_WRONG_FILL),
        ("Light grey row",   "Only one model got this ticker wrong",               ONE_WRONG_FILL),
        ("Green sentiment",  "Positive label",                                     SENTIMENT_FILLS["positive"]),
        ("Red sentiment",    "Negative label",                                     SENTIMENT_FILLS["negative"]),
        ("Yellow sentiment", "Neutral label",                                      SENTIMENT_FILLS["neutral"]),
        ("Dark grey cell",   "Model returned None / skipped (no event identified)", NONE_FILL),
        ("Pale green cell",  "Ground truth columns",                               GT_FILL),
    ]
    for i, (name, desc, fill) in enumerate(legend):
        row = len(summary_rows) + 3 + i
        nc = ws2.cell(row=row, column=1, value=name)
        nc.fill = fill
        nc.font = NORMAL_FONT
        nc.border = THIN_BORDER
        dc = ws2.cell(row=row, column=2, value=desc)
        dc.font = NORMAL_FONT
        dc.border = THIN_BORDER
        ws2.column_dimensions["B"].width = max(ws2.column_dimensions["B"].width, len(desc) * 0.9)

    ws2.column_dimensions["B"].width = 55

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"  {len(both_ids)} posts | {total_ticker_rows} ticker rows")


if __name__ == "__main__":
    main()
